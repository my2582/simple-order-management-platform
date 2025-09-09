#!/usr/bin/env python3
"""
Create sample output Excel file with IBKR standard format
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

def create_sample_output():
    """Create sample output Excel file following IBKR standards"""
    
    # Create sample data for Summary sheet
    summary_data = {
        'Account ID': ['DU123456', 'DU789012', 'Total'],
        'Net Liquidation Value': [1000000.00, 500000.00, 1500000.00],
        'Securities Gross Position Value': [950000.00, 480000.00, 1430000.00],
        'Cash': [50000.00, 20000.00, 70000.00],
        'Gross/NLV': [0.95, 0.96, 0.953],
        'Cash %': [0.05, 0.04, 0.047],
        'Available Funds': [52000.00, 21000.00, 73000.00],
        'Excess Liquidity': [51000.00, 20500.00, 71500.00]
    }
    
    # Create sample data for Matrix sheet
    matrix_data = {
        'Account ID': ['DU123456', 'DU789012'],
        'Net Liquidation Value': [1000000.00, 500000.00],
        'Securities Gross Position Value': [950000.00, 480000.00],
        'Cash': [50000.00, 20000.00],
        'Gross/NLV': [0.95, 0.96],
        'Cash %': [0.05, 0.04],
        'SPMO': [0.33, 0.32],  # Sample position weights
        'SMH': [0.32, 0.33],
        'IAU': [0.30, 0.31],
        'VTI': [0.00, 0.04],   # Some accounts may not have all positions
        'BND': [0.00, 0.00]
    }
    
    # Create Excel file
    output_path = './data/sample_output/sample_output_portfolio_positions_20250906_230000.xlsx'
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Create Summary sheet
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Create Matrix sheet
        matrix_df = pd.DataFrame(matrix_data)
        matrix_df.to_excel(writer, sheet_name='Matrix', index=False)
        
        # Get workbook and format sheets
        workbook = writer.book
        
        # Format Summary sheet
        summary_sheet = workbook['Summary']
        format_sheet(summary_sheet, summary_df)
        
        # Format Matrix sheet
        matrix_sheet = workbook['Matrix']
        format_sheet(matrix_sheet, matrix_df)
        
        # Add metadata sheet
        metadata_data = {
            'Field': ['Export Date', 'Data Source', 'Portfolio Count', 'IBKR Connection', 'Asset Classes'],
            'Value': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Interactive Brokers API',
                '2 accounts',
                'Gateway 4002',
                'Equity ETFs, Fixed Income, Commodities'
            ]
        }
        metadata_df = pd.DataFrame(metadata_data)
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        metadata_sheet = workbook['Metadata']
        format_sheet(metadata_sheet, metadata_df)

def format_sheet(sheet, df):
    """Apply standard formatting to Excel sheet"""
    
    # Header formatting
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Apply header formatting
    for col in range(1, len(df.columns) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

if __name__ == '__main__':
    create_sample_output()
    print("✅ Sample output file created: ./data/sample_output/sample_output_portfolio_positions_20250906_230000.xlsx")