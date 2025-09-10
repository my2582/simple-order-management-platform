"""Holdings table exporter for IBKR standard format."""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional, Any, List
import logging
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..models.portfolio import MultiAccountPortfolio
from ..models.universe import universe_manager, get_asset_class, get_instrument_info
from ..config.loader import config_loader
from ..utils.exceptions import SimpleOrderManagementPlatformError

logger = logging.getLogger(__name__)


class HoldingsExporter:
    """Exporter for IBKR standard holdings table format."""
    
    def __init__(self):
        """Initialize the exporter."""
        self.universe_manager = universe_manager
    
    def export_holdings_table(
        self,
        multi_portfolio: MultiAccountPortfolio,
        output_filename: Optional[str] = None,
        portfolio_service: Optional[Any] = None,
        account_aliases: Optional[Dict[str, str]] = None
    ) -> Path:
        """
        Export holdings data in IBKR standard holdings table format.
        
        Args:
            multi_portfolio: MultiAccountPortfolio object with snapshots
            output_filename: Custom output filename (optional)
            portfolio_service: PortfolioService for getting account aliases
            
        Returns:
            Path to the exported Excel file
        """
        if not multi_portfolio.snapshots:
            raise SimpleOrderManagementPlatformError("No portfolio snapshots to export")
        
        # Generate output filename if not provided
        if output_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"holdings_table_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'
        
        # Get output directory
        app_config = config_loader.load_app_config()
        output_dir = Path(app_config.app.get("directories", {}).get("output_dir", "./data/output"))
        output_dir.mkdir(parents=True, exist_ok=True)
        
        output_path = output_dir / output_filename
        
        logger.info(f"Exporting holdings table to: {output_path}")
        
        # Create Excel file with holdings table
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # Create Holdings sheet
            holdings_df = self._create_holdings_sheet(multi_portfolio, portfolio_service, account_aliases)
            holdings_df.to_excel(writer, sheet_name='Holdings', index=False)
            
            # Format the sheet
            workbook = writer.book
            self._format_workbook(workbook)
        
        # Log export summary
        total_accounts = len(multi_portfolio.snapshots)
        total_positions = sum(len([p for p in s.positions if p.position != 0]) for s in multi_portfolio.snapshots)
        
        logger.info(f"Holdings table export completed: {total_accounts} accounts, "
                   f"{total_positions} total positions exported to {output_path.name}")
        
        return output_path
    
    def _create_holdings_sheet(self, multi_portfolio: MultiAccountPortfolio, portfolio_service: Optional[Any] = None, account_aliases: Optional[Dict[str, str]] = None) -> pd.DataFrame:
        """Create holdings table sheet matching IBKR standard format."""
        
        holdings_data = []
        
        for snapshot in multi_portfolio.snapshots:
            # Get Account Alias from pre-fetched data or via ib_insync (fallback)
            account_alias = None
            if account_aliases and snapshot.account_id in account_aliases:
                account_alias = account_aliases[snapshot.account_id]
            elif portfolio_service:
                try:
                    account_alias = portfolio_service.get_account_alias(snapshot.account_id)
                except Exception as e:
                    logger.debug(f"Could not get account alias for {snapshot.account_id}: {e}")
                    account_alias = None
            
            # Process each position (include zero positions to match IBKR format)
            for position in snapshot.positions:
                
                # Get instrument information
                instrument_info = get_instrument_info(position.symbol)
                
                # Calculate Delta Dollars (approximation: Market Value for equities)
                delta_dollars = 0
                if position.market_value and position.sec_type == 'STK':
                    delta_dollars = float(position.market_value)
                
                # Determine Liquidate Last (placeholder logic)
                liquidate_last = "No"  # Default value, can be enhanced with specific business logic
                
                # Convert values to appropriate formats
                market_price = float(position.market_price) if position.market_price else 0
                market_value = float(position.market_value) if position.market_value else 0
                average_price = float(position.avg_cost) if position.avg_cost else 0
                unrealized_pnl = float(position.unrealized_pnl) if position.unrealized_pnl else 0
                realized_pnl = float(position.realized_pnl) if position.realized_pnl else 0
                position_qty = float(position.position) if position.position else 0
                
                # Handle null values for zero positions (match IBKR format)
                if position_qty == 0:
                    market_value = 0
                    average_price = 0
                    unrealized_pnl = None  # Show as null for zero positions
                    delta_dollars = 0
                
                holdings_data.append({
                    'Account Number': snapshot.account_id,
                    'Account Alias': account_alias or '',
                    'Financial Instrument Description': position.symbol,
                    'Exchange': position.exchange or 'SMART',
                    'Position': position_qty,
                    'Currency': position.currency,
                    'Market Price': market_price,
                    'Market Value': market_value,
                    'Average Price': average_price,
                    'Unrealized P&L': unrealized_pnl,
                    'Realized P&L': realized_pnl,
                    'Liquidate Last': liquidate_last,
                    'Security Type': position.sec_type,
                    'Delta Dollars': delta_dollars
                })
        
        return pd.DataFrame(holdings_data)
    
    def _format_workbook(self, workbook: openpyxl.Workbook) -> None:
        """Apply formatting to the holdings table workbook."""
        
        # Define color scheme
        header_blue = '2E5090'
        light_blue = 'E8F1FF'
        white = 'FFFFFF'
        
        # Define fonts
        header_font = Font(name='Calibri', size=11, bold=True, color=white)
        data_font = Font(name='Calibri', size=10, color='333333')
        
        # Define fills
        header_fill = PatternFill(start_color=header_blue, end_color=header_blue, fill_type='solid')
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            self._format_holdings_sheet(sheet, header_font, data_font, header_fill, thin_border)
    
    def _format_holdings_sheet(self, sheet, header_font, data_font, header_fill, thin_border):
        """Apply formatting to Holdings sheet."""
        
        if sheet.max_row == 0:
            return
        
        # Set freeze panes at A2 (freeze header row)
        sheet.freeze_panes = 'A2'
        
        # Format headers (row 1)
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Format data rows
        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                # Apply specific formatting based on column
                header_cell = sheet.cell(row=1, column=col)
                if header_cell.value and isinstance(cell.value, (int, float)):
                    header_text = str(header_cell.value)
                    
                    # Position column - decimal format
                    if 'Position' in header_text:
                        cell.number_format = '0.0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Price columns - currency format
                    elif any(term in header_text for term in ['Price', 'Value', 'P&L', 'Delta']):
                        if cell.value == 0 and 'P&L' in header_text:
                            # Show 0 for P&L
                            cell.number_format = '0'
                        else:
                            cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Handle null values for Unrealized P&L
                    elif cell.value is None and 'Unrealized P&L' in header_text:
                        cell.value = 'null'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Text alignment for non-numeric columns
                elif col <= 4 or col >= 12:  # Account, Alias, Symbol, Exchange, Liquidate Last, Security Type
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Set column widths for better readability
        column_widths = {
            'A': 15,  # Account Number
            'B': 20,  # Account Alias
            'C': 12,  # Financial Instrument Description
            'D': 12,  # Exchange
            'E': 10,  # Position
            'F': 10,  # Currency
            'G': 15,  # Market Price
            'H': 15,  # Market Value
            'I': 15,  # Average Price
            'J': 15,  # Unrealized P&L
            'K': 15,  # Realized P&L
            'L': 12,  # Liquidate Last
            'M': 12,  # Security Type
            'N': 15,  # Delta Dollars
        }
        
        for col_letter, width in column_widths.items():
            if col_letter <= chr(ord('A') + sheet.max_column - 1):
                sheet.column_dimensions[col_letter].width = width
        
        # Add alternating row colors for better readability
        for row in range(2, sheet.max_row + 1):
            is_even_row = (row % 2 == 0)
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if is_even_row:
                    cell.fill = PatternFill(start_color='FAFBFC', end_color='FAFBFC', fill_type='solid')


# Global instance for easy access
holdings_exporter = HoldingsExporter()


def export_holdings_table(
    multi_portfolio: MultiAccountPortfolio,
    output_filename: Optional[str] = None,
    portfolio_service: Optional[Any] = None,
    account_aliases: Optional[Dict[str, str]] = None
) -> Path:
    """Convenience function to export holdings table."""
    return holdings_exporter.export_holdings_table(
        multi_portfolio=multi_portfolio,
        output_filename=output_filename,
        portfolio_service=portfolio_service,
        account_aliases=account_aliases
    )