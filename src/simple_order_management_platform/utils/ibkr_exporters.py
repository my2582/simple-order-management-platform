"""IBKR standard format exporters for portfolio data."""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional, Any, List
import logging
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..models.portfolio import MultiAccountPortfolio
from ..models.universe import universe_manager, get_asset_class, get_instrument_info
from ..config.loader import config_loader
from ..utils.exceptions import SimpleOrderManagementPlatformError

logger = logging.getLogger(__name__)


class IBKRStandardExporter:
    """Exporter for IBKR standard format portfolio reports."""
    
    def __init__(self):
        """Initialize the exporter."""
        self.universe_manager = universe_manager
    
    def export_portfolio_report(
        self,
        multi_portfolio: MultiAccountPortfolio,
        output_filename: Optional[str] = None,
        include_metadata: bool = True
    ) -> Path:
        """
        Export portfolio data in IBKR standard format.
        
        Args:
            multi_portfolio: MultiAccountPortfolio object with snapshots
            output_filename: Custom output filename (optional)
            include_metadata: Whether to include metadata sheet
            
        Returns:
            Path to the exported Excel file
        """
        if not multi_portfolio.snapshots:
            raise SimpleOrderManagementPlatformError("No portfolio snapshots to export")
        
        # Generate output filename if not provided
        if output_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"portfolio_positions_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'
        
        # Get output directory
        app_config = config_loader.load_app_config()
        output_dir = Path(app_config.app.get("directories", {}).get("output_dir", "./data/output"))
        output_dir.mkdir(parents=True, exist_ok=True)
        
        output_path = output_dir / output_filename
        
        logger.info(f"Exporting portfolio report to: {output_path}")
        
        # Create Excel file with IBKR standard format
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # Create Summary sheet (IBKR account summary format)
            summary_df = self._create_summary_sheet(multi_portfolio)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Create Matrix sheet (positions matrix with asset classes) - shows weights
            matrix_df = self._create_matrix_sheet(multi_portfolio)
            matrix_df.to_excel(writer, sheet_name='Matrix', index=False)
            
            # Create Amt_Matrix sheet (same format as Matrix but with base currency amounts)
            amt_matrix_df = self._create_amt_matrix_sheet(multi_portfolio)
            amt_matrix_df.to_excel(writer, sheet_name='Amt_Matrix', index=False)
            
            # Create metadata sheet if requested
            if include_metadata:
                metadata_df = self._create_metadata_sheet(multi_portfolio)
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
            
            # Format all sheets
            workbook = writer.book
            self._format_workbook(workbook)
        
        # Log export summary
        total_accounts = len(multi_portfolio.snapshots)
        total_positions = sum(len([p for p in s.positions if p.position != 0]) for s in multi_portfolio.snapshots)
        
        logger.info(f"IBKR portfolio export completed: {total_accounts} accounts, "
                   f"{total_positions} total positions exported to {output_path.name}")
        
        return output_path
    
    def _create_summary_sheet(self, multi_portfolio: MultiAccountPortfolio) -> pd.DataFrame:
        """Create enhanced IBKR standard summary sheet with better formatting."""
        
        summary_data = []
        total_nlv = 0
        total_gross_position = 0
        total_cash = 0
        
        for snapshot in multi_portfolio.snapshots:
            account_summary = snapshot.get_positions_summary()
            
            # Calculate IBKR standard values
            nlv = float(account_summary['total_value'])  # Net Liquidation Value
            
            # Calculate Securities Gross Position Value using proper IB API value
            cash_value = 0
            gross_position_value = 0
            
            if snapshot.account_summary:
                cash_value = float(snapshot.account_summary.total_cash_value or 0)
                # Use the proper GrossPositionValue from IB API instead of manual calculation
                # This prevents including master account values
                gross_position_value = float(snapshot.account_summary.gross_position_value or 0)
            
            # Calculate ratios
            gross_nlv_ratio = gross_position_value / nlv if nlv != 0 else 0
            cash_percentage = cash_value / nlv if nlv != 0 else 0
            
            # Calculate additional IBKR metrics
            buying_power = nlv * 4  # Approximate buying power (4:1 margin)
            equity_with_loan = nlv - cash_value  # Equity with loan value
            
            # Add to totals
            total_nlv += nlv
            total_gross_position += gross_position_value
            total_cash += cash_value
            
            summary_data.append({
                'Account ID': snapshot.account_id,
                'Net Liquidation Value': nlv,
                'Total Cash Value': cash_value,
                'Securities Gross Position Value': gross_position_value,
                'Gross/NLV': gross_nlv_ratio,
                'Cash %': cash_percentage,
                'Buying Power': buying_power,
                'Equity w/ Loan Value': equity_with_loan,
                'Available Funds': cash_value + (nlv * 0.25),  # More realistic available funds
                'Excess Liquidity': cash_value * 0.95,  # Conservative excess liquidity
                'Unrealized PnL': 0,  # Placeholder - will be enhanced with real data
                'Realized PnL': 0     # Placeholder - will be enhanced with real data
            })
        
        # Add total row if multiple accounts
        if len(summary_data) > 1:
            total_buying_power = total_nlv * 4
            total_equity_with_loan = total_nlv - total_cash
            
            summary_data.append({
                'Account ID': 'TOTAL',
                'Net Liquidation Value': total_nlv,
                'Total Cash Value': total_cash,
                'Securities Gross Position Value': total_gross_position,
                'Gross/NLV': total_gross_position / total_nlv if total_nlv != 0 else 0,
                'Cash %': total_cash / total_nlv if total_nlv != 0 else 0,
                'Buying Power': total_buying_power,
                'Equity w/ Loan Value': total_equity_with_loan,
                'Available Funds': total_cash + (total_nlv * 0.25),
                'Excess Liquidity': total_cash * 0.95,
                'Unrealized PnL': 0,
                'Realized PnL': 0
            })
        
        return pd.DataFrame(summary_data)
    
    def _create_matrix_sheet(self, multi_portfolio: MultiAccountPortfolio) -> pd.DataFrame:
        """Create portfolio matrix sheet matching the requested format exactly."""
        
        # Collect all unique symbols and their metadata
        all_symbols = set()
        symbol_metadata = {}  # symbol -> {asset_class, instrument_name}
        account_data = {}
        
        for snapshot in multi_portfolio.snapshots:
            account_summary = snapshot.get_positions_summary()
            positions_dict = {}
            
            for pos in account_summary['positions']:
                symbol = pos['Symbol']
                all_symbols.add(symbol)
                
                # Get metadata from universe (priority source as requested)
                instrument_info = get_instrument_info(symbol)
                asset_class = get_asset_class(symbol)
                
                # Store metadata for later use
                symbol_metadata[symbol] = {
                    'asset_class': asset_class or 'Unknown',
                    'instrument_name': instrument_info.instrument_name if instrument_info else symbol
                }
                
                positions_dict[symbol] = {
                    'weight': pos['Weight_Pct'],  # Keep as percentage for proper display
                    'market_value': pos.get('Market_Value', 0)
                }
            
            # Calculate account summary values
            nlv = float(account_summary['total_value'])
            cash_value = 0
            if snapshot.account_summary:
                cash_value = float(snapshot.account_summary.total_cash_value or 0)
            
            # Use proper GrossPositionValue from IB API to avoid master account inclusion
            gross_position_value = float(snapshot.account_summary.gross_position_value or 0) if snapshot.account_summary else 0
            
            account_data[snapshot.account_id] = {
                'positions': positions_dict,
                'nlv': nlv,
                'cash': cash_value,
                'gross_position_value': gross_position_value
            }
        
        # Group symbols by asset class for proper layout
        asset_classes = {}
        for symbol in all_symbols:
            asset_class = symbol_metadata[symbol]['asset_class']
            if asset_class not in asset_classes:
                asset_classes[asset_class] = []
            asset_classes[asset_class].append(symbol)
        
        # Sort asset classes and symbols within each class
        sorted_asset_classes = sorted(asset_classes.keys())
        for asset_class in asset_classes:
            asset_classes[asset_class] = sorted(asset_classes[asset_class])
        
        # Build the matrix in the exact format requested
        matrix_rows = []
        
        # Row 1: Portfolio Matrix header with asset class weight and asset weight sections
        remaining_symbols = []
        for asset_class in sorted_asset_classes:
            remaining_symbols.extend(asset_classes[asset_class])
        
        # Calculate section positions dynamically
        portfolio_matrix_cols = 4  # A-D columns
        asset_class_cols = 4  # E-H columns (Total, Equity, Bond, Gold)
        asset_weight_cols = 1 + len(remaining_symbols)  # Total + individual symbols
        
        header_row1 = ['Portfolio Matrix (Base: S$)', '', '', '', 'Asset class weight', '', '', '', 'Asset weight']
        
        # Add empty cells for remaining asset weight columns
        header_row1.extend([''] * (len(remaining_symbols)))  # Individual symbols
        matrix_rows.append(header_row1)
        
        # Row 2: Export Time and instrument names
        export_time = datetime.now().strftime('%-m/%-d/%y %H:%M')
        header_row2 = ['Export Time', export_time, '', '', 'Total', '', '', '']
        
        # Add instrument names for each symbol with Total first
        if remaining_symbols:
            header_row2.append('Total')  # Total for individual weights
            for symbol in remaining_symbols:
                instrument_name = symbol_metadata[symbol]['instrument_name']
                header_row2.append(instrument_name)
        
        matrix_rows.append(header_row2)
        
        # Row 3: Empty row
        empty_row = [''] * len(header_row2)
        matrix_rows.append(empty_row)
        
        # Row 4: Asset class row
        asset_class_row = ['', '', '', '', 'Total', 'Equity', 'Bond', 'Gold']
        
        # Add asset class for individual symbols with Total first
        if remaining_symbols:
            asset_class_row.append('Total')  # Total for individual weights
            for symbol in remaining_symbols:
                asset_class = symbol_metadata[symbol]['asset_class']
                asset_class_row.append(asset_class)
        
        matrix_rows.append(asset_class_row)
        
        # Row 5: Column headers
        column_headers = ['Account', 'Net Liquidation Value (S$)', 'Gross/NLV', 'Cash % (S$)', 'Total', 'Equity', 'Bond', 'Gold']
        
        # Add symbol codes as column headers with Total at the beginning
        if remaining_symbols:
            column_headers.append('Total')  # Total for asset weight section
            column_headers.extend(remaining_symbols)
        matrix_rows.append(column_headers)
        
        # Data rows: One row per account
        for account_id in sorted(account_data.keys()):
            data = account_data[account_id]
            
            # Calculate asset class weights
            equity_weight = 0
            bond_weight = 0 
            gold_weight = 0
            
            for symbol, pos_data in data['positions'].items():
                asset_class = symbol_metadata[symbol]['asset_class'].lower()
                weight_pct = pos_data['weight']
                
                # Debug logging to understand weight calculation issues
                logger.debug(f"Account {account_id}, Symbol {symbol}: weight_pct={weight_pct}, asset_class={asset_class}")
                
                if 'equity' in asset_class or 'etf' in asset_class:
                    equity_weight += weight_pct
                elif 'bond' in asset_class or 'fixed' in asset_class:
                    bond_weight += weight_pct
                elif 'gold' in asset_class or 'commodity' in asset_class:
                    gold_weight += weight_pct
            
            # Build account row with proper number formatting
            # Convert percentages to decimals for Excel percentage formatting
            
            # Calculate asset class total (for E column - first asset class column)
            asset_class_total = (equity_weight + bond_weight + gold_weight) / 100.0
            
            # Calculate individual symbol weights and their total 
            individual_weights = []
            individual_weights_total = 0
            for symbol in remaining_symbols:
                if symbol in data['positions']:
                    weight = data['positions'][symbol]['weight'] / 100.0 if data['positions'][symbol]['weight'] != 0 else 0
                else:
                    weight = 0
                individual_weights.append(weight)
                individual_weights_total += weight
            
            account_row = [
                account_id,
                data['nlv'],  # Will be formatted as #,###.00 in Excel
                (data['gross_position_value'] / data['nlv']) if data['nlv'] != 0 else 0,  # Will be formatted as 0.00
                (data['cash'] / data['nlv']) if data['nlv'] != 0 else 0,  # Will be formatted as 0.00% in Excel
                asset_class_total,  # Total of asset class weights (E column - first in asset class section)
                equity_weight / 100.0 if equity_weight != 0 else 0,
                bond_weight / 100.0 if bond_weight != 0 else 0,
                gold_weight / 100.0 if gold_weight != 0 else 0
            ]
            
            # Add individual symbol weights with total in first asset weight column
            if remaining_symbols:
                account_row.append(individual_weights_total)  # Total in first asset weight column 
                account_row.extend(individual_weights)  # All individual weights
            else:
                account_row.append(0)  # Total is 0 if no symbols
            
            matrix_rows.append(account_row)
        
        # Convert to DataFrame
        # Find the maximum row length to ensure all rows have the same number of columns
        max_cols = max(len(row) for row in matrix_rows)
        
        # Pad shorter rows with empty strings
        for row in matrix_rows:
            while len(row) < max_cols:
                row.append('')
        
        # Create DataFrame without headers (we'll handle headers in formatting)
        df = pd.DataFrame(matrix_rows)
        return df
    
    def _create_amt_matrix_sheet(self, multi_portfolio: MultiAccountPortfolio) -> pd.DataFrame:
        """Create amount matrix sheet - same format as Matrix but showing base currency amounts instead of weights."""
        
        # Collect all unique symbols and their metadata (same as Matrix)
        all_symbols = set()
        symbol_metadata = {}  # symbol -> {asset_class, instrument_name}
        account_data = {}
        
        for snapshot in multi_portfolio.snapshots:
            account_summary = snapshot.get_positions_summary()
            positions_dict = {}
            
            for pos in account_summary['positions']:
                symbol = pos['Symbol']
                all_symbols.add(symbol)
                
                # Get metadata from universe (priority source as requested)
                instrument_info = get_instrument_info(symbol)
                asset_class = get_asset_class(symbol)
                
                # Store metadata for later use
                symbol_metadata[symbol] = {
                    'asset_class': asset_class or 'Unknown',
                    'instrument_name': instrument_info.instrument_name if instrument_info else symbol
                }
                
                positions_dict[symbol] = {
                    'market_value': pos.get('Market_Value', 0)  # Use market value in base currency
                }
            
            # Calculate account summary values
            nlv = float(account_summary['total_value'])
            cash_value = 0
            if snapshot.account_summary:
                cash_value = float(snapshot.account_summary.total_cash_value or 0)
            
            # Use proper GrossPositionValue from IB API
            gross_position_value = float(snapshot.account_summary.gross_position_value or 0) if snapshot.account_summary else 0
            
            account_data[snapshot.account_id] = {
                'positions': positions_dict,
                'nlv': nlv,
                'cash': cash_value,
                'gross_position_value': gross_position_value
            }
        
        # Group symbols by asset class for proper layout
        asset_classes = {}
        for symbol in all_symbols:
            asset_class = symbol_metadata[symbol]['asset_class']
            if asset_class not in asset_classes:
                asset_classes[asset_class] = []
            asset_classes[asset_class].append(symbol)
        
        # Sort asset classes and symbols within each class
        sorted_asset_classes = sorted(asset_classes.keys())
        for asset_class in asset_classes:
            asset_classes[asset_class] = sorted(asset_classes[asset_class])
        
        # Build the matrix in the exact format requested
        matrix_rows = []
        
        # Row 1: Portfolio Matrix header with asset class amount and asset amount sections
        remaining_symbols = []
        for asset_class in sorted_asset_classes:
            remaining_symbols.extend(asset_classes[asset_class])
        
        header_row1 = ['Portfolio Matrix (S$)', '', '', '', 'Asset class amount (S$)', '', '', '', 'Asset amount (S$)']
        
        # Add empty cells for remaining asset amount columns
        header_row1.extend([''] * (len(remaining_symbols)))  # Individual symbols
        matrix_rows.append(header_row1)
        
        # Row 2: Export Time and instrument names
        export_time = datetime.now().strftime('%-m/%-d/%y %H:%M')
        header_row2 = ['Export Time', export_time, '', '', 'Total', '', '', '']
        
        # Add instrument names for each symbol with Total first
        if remaining_symbols:
            header_row2.append('Total')  # Total for individual amounts
            for symbol in remaining_symbols:
                instrument_name = symbol_metadata[symbol]['instrument_name']
                header_row2.append(instrument_name)
        
        matrix_rows.append(header_row2)
        
        # Row 3: Empty row
        empty_row = [''] * len(header_row2)
        matrix_rows.append(empty_row)
        
        # Row 4: Asset class row
        asset_class_row = ['', '', '', '', 'Total', 'Equity', 'Bond', 'Gold']
        
        # Add asset class for individual symbols with Total first
        if remaining_symbols:
            asset_class_row.append('Total')  # Total for individual amounts
            for symbol in remaining_symbols:
                asset_class = symbol_metadata[symbol]['asset_class']
                asset_class_row.append(asset_class)
        
        matrix_rows.append(asset_class_row)
        
        # Row 5: Column headers
        column_headers = ['Account', 'Net Liquidation Value (S$)', 'Gross/NLV', 'Cash % (S$)', 'Total (S$)', 'Equity (S$)', 'Bond (S$)', 'Gold (S$)']
        
        # Add symbol codes as column headers with Total at the beginning
        if remaining_symbols:
            column_headers.append('Total (S$)')  # Total for asset amount section
            column_headers.extend([f'{symbol} (S$)' for symbol in remaining_symbols])
        matrix_rows.append(column_headers)
        
        # Data rows: One row per account
        for account_id in sorted(account_data.keys()):
            data = account_data[account_id]
            
            # Calculate asset class amounts in SGD
            equity_amount = 0
            bond_amount = 0 
            gold_amount = 0
            
            for symbol, pos_data in data['positions'].items():
                asset_class = symbol_metadata[symbol]['asset_class'].lower()
                market_value = pos_data['market_value']
                
                if 'equity' in asset_class or 'etf' in asset_class:
                    equity_amount += market_value
                elif 'bond' in asset_class or 'fixed' in asset_class:
                    bond_amount += market_value
                elif 'gold' in asset_class or 'commodity' in asset_class:
                    gold_amount += market_value
            
            # Calculate individual symbol amounts and their total 
            individual_amounts = []
            individual_amounts_total = 0
            for symbol in remaining_symbols:
                if symbol in data['positions']:
                    amount = data['positions'][symbol]['market_value']
                else:
                    amount = 0
                individual_amounts.append(amount)
                individual_amounts_total += amount
            
            # Build account row
            account_row = [
                account_id,
                data['nlv'],  # Net Liquidation Value in SGD
                (data['gross_position_value'] / data['nlv']) if data['nlv'] != 0 else 0,  # Ratio
                data['cash'],  # Cash amount in SGD
                equity_amount + bond_amount + gold_amount,  # Total asset class amounts
                equity_amount,  # Equity amount in SGD
                bond_amount,    # Bond amount in SGD
                gold_amount     # Gold amount in SGD
            ]
            
            # Add individual symbol amounts with total in first asset amount column
            if remaining_symbols:
                account_row.append(individual_amounts_total)  # Total in first asset amount column 
                account_row.extend(individual_amounts)  # All individual amounts
            else:
                account_row.append(0)  # Total is 0 if no symbols
            
            matrix_rows.append(account_row)
        
        # Convert to DataFrame
        # Find the maximum row length to ensure all rows have the same number of columns
        max_cols = max(len(row) for row in matrix_rows)
        
        # Pad shorter rows with empty strings
        for row in matrix_rows:
            while len(row) < max_cols:
                row.append('')
        
        # Create DataFrame without headers (we'll handle headers in formatting)
        df = pd.DataFrame(matrix_rows)
        return df
    
    def _create_metadata_sheet(self, multi_portfolio: MultiAccountPortfolio) -> pd.DataFrame:
        """Create metadata sheet with export information."""
        
        # Get universe summary
        universe_summary = self.universe_manager.get_universe_summary()
        
        # Get price date information (will be enhanced when market data platform is integrated)
        price_date = datetime.now().strftime('%Y-%m-%d')  # Placeholder
        
        metadata_data = [
            ['Export Information', ''],
            ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Data Source', 'Interactive Brokers API'],
            ['Portfolio Count', len(multi_portfolio.snapshots)],
            ['Price Date Used', price_date],
            ['', ''],
            ['Universe Information', ''],
            ['Total Instruments', universe_summary['total_instruments']],
            ['Asset Classes', universe_summary['total_asset_classes']],
            ['Asset Class Breakdown', ''],
        ]
        
        # Add asset class breakdown
        for asset_class, count in universe_summary['asset_class_breakdown'].items():
            metadata_data.append([f"  - {asset_class}", count])
        
        return pd.DataFrame(metadata_data, columns=['Field', 'Value'])
    
    def _format_workbook(self, workbook: openpyxl.Workbook) -> None:
        """Apply sophisticated formatting to the workbook with freeze panes and proper styling."""
        
        # Define sophisticated color scheme
        primary_blue = '2E5090'      # Deep professional blue
        secondary_blue = '4A6FA5'    # Medium blue
        light_blue = 'E8F1FF'       # Very light blue background
        header_gray = '404040'       # Dark gray for headers
        light_gray = 'F5F5F5'       # Light gray background
        white = 'FFFFFF'
        
        # Define fonts
        header_font = Font(name='Calibri', size=11, bold=True, color=white)
        subheader_font = Font(name='Calibri', size=10, bold=True, color=header_gray)
        data_font = Font(name='Calibri', size=10, color='333333')
        title_font = Font(name='Calibri', size=12, bold=True, color=primary_blue)
        
        # Define fills
        header_fill = PatternFill(start_color=primary_blue, end_color=primary_blue, fill_type='solid')
        subheader_fill = PatternFill(start_color=secondary_blue, end_color=secondary_blue, fill_type='solid')
        light_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type='solid')
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        thick_border = Border(
            left=Side(style='medium', color=primary_blue),
            right=Side(style='medium', color=primary_blue),
            top=Side(style='medium', color=primary_blue),
            bottom=Side(style='medium', color=primary_blue)
        )
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            if sheet_name == 'Matrix':
                self._format_matrix_sheet(sheet, header_font, subheader_font, data_font, title_font, 
                                        header_fill, subheader_fill, light_fill, thin_border, thick_border)
            elif sheet_name == 'Summary':
                self._format_summary_sheet(sheet, header_font, data_font, header_fill, thin_border)
            else:
                self._format_generic_sheet(sheet, header_font, data_font, header_fill, thin_border)
    
    def _format_matrix_sheet(self, sheet, header_font, subheader_font, data_font, title_font, 
                           header_fill, subheader_fill, light_fill, thin_border, thick_border):
        """Apply special formatting to Matrix sheet with freeze panes."""
        
        if sheet.max_row == 0:
            return
            
        # Row 1: Professional section headers with enhanced styling
        # Portfolio Matrix title
        title_cell = sheet.cell(row=1, column=1)
        title_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Asset class weight section header
        asset_class_cell = sheet.cell(row=1, column=5)
        asset_class_cell.value = 'Asset class weight'
        asset_class_cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        asset_class_cell.fill = PatternFill(start_color='4A6FA5', end_color='4A6FA5', fill_type='solid')
        asset_class_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Asset weight section header
        asset_weight_cell = sheet.cell(row=1, column=8)
        asset_weight_cell.value = 'Asset weight'
        asset_weight_cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        asset_weight_cell.fill = PatternFill(start_color='4A6FA5', end_color='4A6FA5', fill_type='solid')
        asset_weight_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells for section headers
        if sheet.max_column >= 9:
            # Merge Portfolio Matrix title
            sheet.merge_cells('A1:D1')
            sheet.merge_cells('E1:H1')  # Asset class weight section (Total, Equity, Bond, Gold)
            
            # Calculate end column for Asset weight section
            end_col = sheet.max_column
            if end_col >= 9:
                end_col_letter = chr(ord('A') + end_col - 1)
                sheet.merge_cells(f'I1:{end_col_letter}1')  # Asset weight section
        
        # Apply styling to merged cells
        for col in range(1, 5):  # Portfolio Matrix section
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(5, 9):  # Asset class weight section (E-H)
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4A6FA5', end_color='4A6FA5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        for col in range(9, sheet.max_column + 1):  # Asset weight section (I onwards)
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4A6FA5', end_color='4A6FA5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # Row 2: Export time with enhanced styling
        export_time_cell = sheet.cell(row=2, column=1)
        export_time_cell.font = Font(name='Calibri', size=10, bold=True, color='2E5090')
        
        export_value_cell = sheet.cell(row=2, column=2)
        export_value_cell.font = Font(name='Calibri', size=10, color='333333')
        
        # Row 2: Instrument names with subtle styling
        for col in range(9, sheet.max_column + 1):
            cell = sheet.cell(row=2, column=col)
            if cell.value:
                cell.font = Font(name='Calibri', size=9, color='666666')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 4: Asset class labels with enhanced styling
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=4, column=col)
            if cell.value:
                cell.font = Font(name='Calibri', size=10, bold=True, color='2E5090')
                cell.fill = PatternFill(start_color='F0F5FF', end_color='F0F5FF', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 5: Column headers (main data headers)
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=5, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thick_border
        
        # Data rows (6 onwards)
        for row in range(6, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                # Apply sophisticated number formatting based on column content
                header_cell = sheet.cell(row=5, column=col)
                if header_cell.value and isinstance(cell.value, (int, float)):
                    header_text = str(header_cell.value)
                    
                    # Net Liquidation Value: Currency format with commas
                    if 'Net Liquidation Value' in header_text:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Gross/NLV ratio: Plain decimal format  
                    elif 'Gross/NLV' in header_text:
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Cash percentage: Percentage format
                    elif 'Cash %' in header_text:
                        cell.number_format = '0.00%'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Asset class weights (Total, Equity, Bond, Gold): Percentage format (0.00%)
                    elif any(term in header_text for term in ['Equity', 'Bond', 'Gold', 'Total']) and col >= 5 and col <= 8:
                        cell.number_format = '0.00%'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        # Special styling for Total column in asset class section
                        if 'Total' in header_text and col == 5:
                            if cell.value > 0:
                                cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                                cell.font = Font(name='Calibri', size=10, bold=True, color='155724')
                        # Subtle highlighting for non-zero asset class weights
                        elif cell.value > 0:
                            cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                            cell.font = Font(name='Calibri', size=10, bold=True, color='2E5090')
                    
                    # Individual symbol weights: Percentage format (0.00%) with highlighting
                    elif col > 8:  # Individual symbol weights (including Total column for asset weights)
                        cell.number_format = '0.00%'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        # Special styling for Total column in asset weight section
                        if col == 9 and 'Total' in str(sheet.cell(row=5, column=col).value):
                            if cell.value > 0:
                                cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                                cell.font = Font(name='Calibri', size=10, bold=True, color='856404')
                        # Highlight non-zero individual positions with subtle background
                        elif cell.value > 0:
                            cell.fill = PatternFill(start_color='FFF2E8', end_color='FFF2E8', fill_type='solid')
                            cell.font = Font(name='Calibri', size=10, bold=True, color='B8860B')
                
                # Account name column - left align
                elif col == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font = Font(name='Calibri', size=10, bold=True, color='2E5090')
        
        # Set optimized column widths for professional appearance
        column_widths = {
            'A': 15,  # Account - wider for account names
            'B': 20,  # Net Liquidation Value - wider for currency values
            'C': 10,  # Gross/NLV - compact for ratios
            'D': 10,  # Cash % - compact for percentages
            'E': 12,  # Total (Asset class) - slightly wider for totals
            'F': 10,  # Equity - compact for percentages
            'G': 10,  # Bond - compact for percentages
            'H': 10,  # Gold - compact for percentages
            'I': 12,  # Total (Asset weight) - slightly wider for totals
        }
        
        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width
        
        # Set width for individual symbol columns (make them more readable)
        for col in range(10, sheet.max_column + 1):
            col_letter = chr(ord('A') + col - 1)
            sheet.column_dimensions[col_letter].width = 10  # Increased from 8 to 10 for better readability
        
        # Freeze panes: Freeze first 4 columns and first 5 rows
        sheet.freeze_panes = 'E6'
        
        # Add professional alternating row colors for data rows with better logic
        for row in range(6, sheet.max_row + 1):
            is_even_row = (row % 2 == 0)
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                
                # Check if cell already has conditional formatting (non-zero values)
                has_conditional_formatting = False
                if hasattr(cell, 'fill') and cell.fill and hasattr(cell.fill, 'start_color'):
                    current_color = cell.fill.start_color.rgb
                    if current_color and current_color not in [None, 'FFFFFF', '00FFFFFF', '00000000']:
                        has_conditional_formatting = True
                
                # Apply alternating row color only if no conditional formatting exists
                if not has_conditional_formatting and is_even_row:
                    cell.fill = PatternFill(start_color='FAFBFC', end_color='FAFBFC', fill_type='solid')
                elif not has_conditional_formatting:
                    # Ensure odd rows have clean white background
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                # Add subtle hover-like effect for account column
                if col == 1 and cell.value:  # Account ID column
                    cell.fill = PatternFill(start_color='E8F1FF', end_color='E8F1FF', fill_type='solid')
                    cell.font = Font(name='Calibri', size=10, bold=True, color='2E5090')
        
        # Add separator lines between major sections
        # Vertical separator between account info and asset class weights (column D)
        for row in range(5, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=4)  # Column D
            cell.border = Border(
                left=cell.border.left,
                right=Side(style='medium', color='2E5090'),
                top=cell.border.top,
                bottom=cell.border.bottom
            )
        
        # Vertical separator between asset class weights and individual weights (column H) 
        for row in range(5, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=8)  # Column H (after Gold column)
            cell.border = Border(
                left=cell.border.left,
                right=Side(style='medium', color='2E5090'),
                top=cell.border.top,
                bottom=cell.border.bottom
            )
    
    def _format_summary_sheet(self, sheet, header_font, data_font, header_fill, thin_border):
        """Apply formatting to Summary sheet."""
        
        if sheet.max_row == 0:
            return
            
        # Format headers
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Format data rows
        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                # Number formatting based on column headers
                header_cell = sheet.cell(row=1, column=col)
                if header_cell.value and isinstance(cell.value, (int, float)):
                    header_text = str(header_cell.value)
                    
                    if any(term in header_text for term in ['Value', 'Cash', 'Funds', 'Liquidity']):
                        cell.number_format = '#,##0.00'
                    elif '%' in header_text or 'NLV' in header_text:
                        cell.number_format = '0.00%'
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 20)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_generic_sheet(self, sheet, header_font, data_font, header_fill, thin_border):
        """Apply generic formatting to other sheets."""
        
        if sheet.max_row == 0:
            return
            
        # Format headers
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths and add borders
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                    cell.border = thin_border
                    if cell.row > 1:
                        cell.font = data_font
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            sheet.column_dimensions[column_letter].width = adjusted_width


# Global instance for easy access
ibkr_exporter = IBKRStandardExporter()


def export_ibkr_portfolio_report(
    multi_portfolio: MultiAccountPortfolio,
    output_filename: Optional[str] = None,
    include_metadata: bool = True
) -> Path:
    """Convenience function to export IBKR standard portfolio report."""
    return ibkr_exporter.export_portfolio_report(
        multi_portfolio=multi_portfolio,
        output_filename=output_filename,
        include_metadata=include_metadata
    )